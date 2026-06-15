import database
import extractor
import json
import re
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import io
import base64

def get_year_num(year_str):
    m = re.search(r'\d+', year_str)
    return int(m.group(0)) if m else 0

def process_metrics(kpi_records):
    groups = {}
    for r in kpi_records:
        kpi = r['kpi_name']
        if kpi not in groups: groups[kpi] = []
        groups[kpi].append(dict(r))

    all_processed = []
    for kpi, records in groups.items():
        sorted_records = sorted(records, key=lambda x: get_year_num(x['fiscal_year']))
        for i in range(len(sorted_records)):
            sorted_records[i]['yoy_growth'] = None
            if i > 0:
                prev, curr = sorted_records[i-1], sorted_records[i]
                if get_year_num(curr['fiscal_year']) - get_year_num(prev['fiscal_year']) == 1 and prev['kpi_value_numeric'] != 0:
                    sorted_records[i]['yoy_growth'] = round(((curr['kpi_value_numeric'] - prev['kpi_value_numeric']) / abs(prev['kpi_value_numeric'])) * 100, 2)

        cagr = None
        if len(sorted_records) >= 2:
            first, last = sorted_records[0], sorted_records[-1]
            n = get_year_num(last['fiscal_year']) - get_year_num(first['fiscal_year'])
            if n >= 1 and first['kpi_value_numeric'] > 0 and last['kpi_value_numeric'] > 0:
                cagr = round(((last['kpi_value_numeric'] / first['kpi_value_numeric']) ** (1/n) - 1) * 100, 2)

        for r in sorted_records:
            r['cagr'] = cagr
            all_processed.append(r)
    return all_processed

def init_app():
    database.init_db()
    return "OK"

def add_doc(filename, path):
    return database.add_document(filename, path)

def list_docs():
    return json.dumps(database.get_documents())

def delete_doc(doc_id):
    database.delete_document(doc_id)
    return "OK"

def run_extraction(doc_id, kpis_json, custom_json):
    doc = database.get_document(doc_id)
    kpis = json.loads(kpis_json)
    custom = json.loads(custom_json)
    database.clear_kpis_for_document(doc_id)
    results = extractor.run_extraction(doc['filepath'], kpis, custom)
    for r in results:
        database.save_extracted_kpi(doc_id, r['kpi_name'], r['kpi_value_raw'], r['kpi_value_numeric'],
                                   r['fiscal_year'], r['page_number'], r['source_text'], r['confidence'], r['is_custom'])
    return get_results(doc_id)

def get_results(doc_id):
    kpis = database.get_extracted_kpis(doc_id)
    return json.dumps(process_metrics(kpis))

def export_excel(doc_id):
    kpis = process_metrics(database.get_extracted_kpis(doc_id))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Financial Analysis"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    # Headers
    headers = ["KPI Name", "Fiscal Year", "Value (Raw)", "Value (Numeric)", "Confidence (%)", "Page", "YoY Growth (%)", "CAGR (%)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # Data
    for row_idx, kpi in enumerate(kpis, 2):
        ws.cell(row=row_idx, column=1, value=kpi['kpi_name'])
        ws.cell(row=row_idx, column=2, value=kpi['fiscal_year'])
        ws.cell(row=row_idx, column=3, value=kpi['kpi_value_raw'])
        ws.cell(row=row_idx, column=4, value=kpi['kpi_value_numeric'])
        ws.cell(row=row_idx, column=5, value=kpi['confidence'])
        ws.cell(row=row_idx, column=6, value=kpi['page_number'])
        ws.cell(row=row_idx, column=7, value=kpi.get('yoy_growth'))
        ws.cell(row=row_idx, column=8, value=kpi.get('cagr'))

    buffer = io.BytesIO()
    wb.save(buffer)
    return base64.b64encode(buffer.getvalue()).decode('utf-8')
